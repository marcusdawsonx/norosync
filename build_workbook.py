"""Build the NoroSync workbook.

STYLING IS PART OF THE CONTRACT. Every colour, fill and number format below was
read out of the 24.07.2026 workbook rather than invented, because the visual
grammar carries meaning:

  * Channel colour bands - Shopify grey, Meta blue, Google green, AppLovin sand,
    Klaviyo lilac - let you scan across a wide row and know which platform a
    number belongs to without reading the header.
  * A rust header (9C5148) instead of navy marks a column that did NOT come from
    this tab's own source: the all-traffic columns on Landing Pages, and the
    three columns copied down from Landing Pages onto Creative Coverage.
  * Numbers display as $1,234.56, 2.48x and 3.09%, while the underlying cell
    stays a plain number so it still sums and sorts.
  * Paused creatives are greyed out rather than removed, so a dead winner is
    still visible in context.

Formula policy: any cell whose inputs are visible in the sheet is a formula, so
the workbook recalculates when edited. ROAS columns are values because their
numerators are not part of the frozen layout.
"""
import sys
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/root/noro-sync/pipeline")

FONT = "Arial"
NAVY = "1F3864"
RUST = "9C5148"          # header for borrowed / all-traffic columns
TOTAL_FILL = "D9E2F3"
SIGNAL_ON = "B3541E"
SIGNAL_OFF = "8A8A8A"
MUTED = "9AA4AE"         # paused creatives

SHOPIFY = "F2F2F2"
META = "DEEBF7"
GOOGLE = "EAF3EE"
APPLOVIN = "F8F1E9"
KLAVIYO = "F0EAF6"
BORROWED = "FDF0EE"
CREATIVE_BG = "F7F9FC"

MONEY = '$#,##0.00'
MONEY0 = '$#,##0'
ROAS = '0.00"x"'
PCT2 = '0.00%'
PCT1 = '0.0%'
COUNT = '#,##0'
PLAIN = '0.00'


def band(spec, ncols):
    """spec: list of (first_col, last_col, fill). 1-based, inclusive."""
    out = {}
    for a, b, fill in spec:
        for c in range(a, b + 1):
            out[c] = fill
    return [out.get(c) for c in range(1, ncols + 1)]


SNAPSHOT_BANDS = band([(1, 4, SHOPIFY), (5, 6, META), (7, 8, GOOGLE),
                       (9, 10, APPLOVIN), (11, 11, KLAVIYO)], 11)
SNAPSHOT_FMT = {2: MONEY, 3: MONEY, 4: ROAS, 5: MONEY, 6: ROAS, 7: MONEY,
                8: ROAS, 9: MONEY, 10: ROAS, 11: MONEY}

DAILY_BANDS = band([(1, 6, SHOPIFY), (7, 15, META), (16, 24, GOOGLE),
                    (25, 30, APPLOVIN)], 30)
DAILY_FMT = {2: MONEY, 3: COUNT, 4: MONEY, 5: MONEY, 6: MONEY,
             7: MONEY, 8: ROAS, 9: MONEY, 10: ROAS, 11: MONEY, 12: ROAS,
             13: MONEY, 14: PCT2, 15: PCT2,
             16: MONEY, 17: ROAS, 18: MONEY, 19: ROAS, 20: MONEY, 21: ROAS,
             22: MONEY, 23: PCT2, 24: PCT2,
             25: MONEY, 26: ROAS, 27: ROAS, 28: MONEY, 29: PCT2, 30: PCT2}

LP_FMT = {3: PCT1, 4: COUNT, 5: COUNT, 6: MONEY0, 7: MONEY, 8: PCT2, 9: PCT2,
          10: MONEY0, 11: MONEY0, 12: MONEY, 13: MONEY, 14: ROAS}
LP_RUST = {5, 9, 11, 13}
LP_BORROWED_FILL = {5, 9, 11, 13}

CC_FMT = {4: COUNT, 5: ROAS, 6: MONEY, 7: COUNT, 8: COUNT, 9: MONEY0,
          10: COUNT, 11: ROAS, 12: PCT1, 13: COUNT}
CC_RUST = {4, 5, 6}
CC_BORROWED_FILL = {4, 5, 6}

CR_FMT = {6: COUNT, 7: MONEY0, 8: PCT2, 9: PLAIN, 10: PCT2, 11: PCT2,
          12: COUNT, 13: MONEY0, 14: ROAS}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def write_tab(ws, rows, header_row, fmts, bands=None, rust=(), borrowed=(),
              formula_map=None, total_rows=(), muted_rows=(), signal_cols=(),
              base_fill=None, widths=None):
    thin = Side(style="thin", color="D9D9D9")
    ncols = len(rows[header_row - 1])
    for i, row in enumerate(rows, start=1):
        is_total = i in total_rows
        for j in range(1, ncols + 1):
            raw = row[j - 1] if j <= len(row) else ""
            c = ws.cell(row=i, column=j)
            f = (formula_map or {}).get((i, j))
            if f is not None:
                c.value = f
            else:
                v = _num(raw)
                c.value = raw if v is None else v
            c.font = Font(name=FONT, size=10)
            c.border = Border(bottom=thin)

            if i == 1:
                c.font = Font(name=FONT, size=12, bold=True)
            elif i == header_row:
                c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=RUST if j in rust else NAVY)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                if j in fmts:
                    c.number_format = fmts[j]
            elif is_total:
                c.font = Font(name=FONT, size=10, bold=True)
                c.fill = PatternFill("solid", fgColor=TOTAL_FILL)
                if j in fmts:
                    c.number_format = fmts[j]
            elif str(row[0]).strip() in ("Period", "Week"):
                # Sub-table headers for the monthly and weekly blocks. The old
                # report left these as bare text, so the blocks read as stray
                # rows rather than as their own small tables.
                c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=NAVY)
            elif i > header_row:
                # Channel banding applies to the data block only. Below the
                # TOTAL row sit the monthly/weekly sub-tables and the notes,
                # which are not per-channel and must not inherit the colours.
                in_block = i < min(total_rows) if total_rows else True
                fill = (BORROWED if j in borrowed
                        else ((bands[j - 1] if bands else base_fill)
                              if in_block else None))
                if fill:
                    c.fill = PatternFill("solid", fgColor=fill)
                if j in fmts:
                    c.number_format = fmts[j]
                if i in muted_rows:
                    c.font = Font(name=FONT, size=10, color=MUTED)
                if j in signal_cols:
                    on = str(raw).strip() not in ("", "–", "-")
                    c.font = Font(name=FONT, size=9, bold=on,
                                  color=SIGNAL_ON if on else SIGNAL_OFF)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    for j in range(1, ncols + 1):
        w = max((len(str(r[j - 1])) for r in rows[:60] if j <= len(r)), default=10)
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 10), 40)
    for col, wdt in (widths or {}).items():
        ws.column_dimensions[col].width = wdt
    return ws


def _colour_scale(ws, col_letter, first, last):
    if last >= first:
        ws.conditional_formatting.add(
            f"{col_letter}{first}:{col_letter}{last}",
            ColorScaleRule(start_type="min", start_color="FFF7F3EC",
                           mid_type="percentile", mid_value=50, mid_color="FFF3D5B5",
                           end_type="max", end_color="FFE8A87C"))


CONVERSION_LAG = [
    ["Time known before ordering", "% of respondents", "Cumulative %"],
    ["Less than 1 day", 0.2597, 0.2597], ["1 - 7 days", 0.1898, 0.4495],
    ["8 - 14 days", 0.1134, 0.5629], ["15 - 30 days", 0.1269, 0.6898],
    ["1 - 6 months", 0.1995, 0.8893], ["6 - 12 months", 0.0708, 0.9601],
    ["More than 1 year", 0.0398, 0.9999],
    ["NOTE: static customer-survey data. It does not change between runs and is "
     "not re-pulled; source is the post-purchase survey."],
]


def _snapshot_formulas(first, last, total_r):
    fm = {}
    for r in range(first, last + 1):
        fm[(r, 3)] = f"=E{r}+G{r}+I{r}"
        fm[(r, 4)] = f"=IF(C{r}=0,\"\",B{r}/C{r})"
    for col in (2, 5, 7, 9, 11):
        L = get_column_letter(col)
        fm[(total_r, col)] = f"=SUM({L}{first}:{L}{last})"
    fm[(total_r, 3)] = f"=E{total_r}+G{total_r}+I{total_r}"
    fm[(total_r, 4)] = f"=IF(C{total_r}=0,\"\",B{total_r}/C{total_r})"
    return fm


def _daily_formulas(first, last, total_r):
    fm = {}
    for r in range(first, last + 1):
        fm[(r, 4)] = f"=IF(C{r}=0,\"\",B{r}/C{r})"
        fm[(r, 6)] = f"=G{r}+P{r}+Y{r}"
    for col in (2, 3, 5, 7, 9, 11, 16, 18, 20, 25):
        L = get_column_letter(col)
        fm[(total_r, col)] = f"=SUM({L}{first}:{L}{last})"
    fm[(total_r, 4)] = f"=IF(C{total_r}=0,\"\",B{total_r}/C{total_r})"
    fm[(total_r, 6)] = f"=G{total_r}+P{total_r}+Y{total_r}"
    return fm


def _add_core(wb, sn, dd):
    first, last, total_r = 3, 32, 33
    ws = wb.create_sheet("Daily Snapshot")
    write_tab(ws, sn, 2, SNAPSHOT_FMT, bands=SNAPSHOT_BANDS,
              formula_map=_snapshot_formulas(first, last, total_r),
              total_rows={total_r}, widths={"A": 16})
    ws = wb.create_sheet("Daily Data")
    write_tab(ws, dd, 2, DAILY_FMT, bands=DAILY_BANDS,
              formula_map=_daily_formulas(first, last, total_r),
              total_rows={total_r}, widths={"A": 12})


def _total_row_index(rows):
    for i, r in enumerate(rows, start=1):
        if str(r[0]).startswith("TOTAL"):
            return i
    return -1


def build(sn, dd, path):
    wb = Workbook()
    wb.remove(wb.active)
    _add_core(wb, sn, dd)
    wb.save(path)
    return path


def build_full(sn, dd, lp, cc, cr, path):
    wb = Workbook()
    wb.remove(wb.active)
    _add_core(wb, sn, dd)

    ws = wb.create_sheet("Meta Landing Pages")
    write_tab(ws, lp, 2, LP_FMT, rust=LP_RUST, borrowed=LP_BORROWED_FILL,
              total_rows={_total_row_index(lp)}, signal_cols={15},
              widths={"A": 38, "O": 22})

    ws = wb.create_sheet("Creative Coverage")
    write_tab(ws, cc, 2, CC_FMT, rust=CC_RUST, borrowed=CC_BORROWED_FILL,
              total_rows={_total_row_index(cc)}, signal_cols={3, 14},
              widths={"A": 38, "C": 14, "N": 24})
    _colour_scale(ws, "I", 3, _total_row_index(cc) - 1)

    ws = wb.create_sheet("Creative Inventory")
    muted = {i for i, r in enumerate(cr, start=1)
             if i > 2 and str(r[2]).strip() not in ("", "ACTIVE", "Status")}
    write_tab(ws, cr, 2, CR_FMT, base_fill=CREATIVE_BG,
              total_rows={_total_row_index(cr)}, muted_rows=muted,
              signal_cols={15}, widths={"A": 40, "B": 28, "O": 19})
    _colour_scale(ws, "G", 3, _total_row_index(cr) - 1)

    ws = wb.create_sheet("Conversion Lag")
    write_tab(ws, CONVERSION_LAG, 1, {2: PCT2, 3: PCT2}, widths={"A": 26})

    wb.save(path)
    return path
